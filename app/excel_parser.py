import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
from app.models import Season, Participant, Event, Result
from app.config import PLACE_POINTS, PARTICIPATION_POINTS

SCHOOL_EVENTS = [
    {"name": "Winter", "emoji": "❄️", "multiplier": 1, "sort_order": 1},
    {"name": "Spring", "emoji": "🌸", "multiplier": 1, "sort_order": 2},
    {"name": "Summer", "emoji": "☀️", "multiplier": 1, "sort_order": 3},
    {"name": "Autumn", "emoji": "🍂", "multiplier": 1, "sort_order": 4},
    {"name": "Final", "emoji": "🔥", "multiplier": 2, "sort_order": 5},
]

SHEET_MAP = {
    "❄️ Winter": "Winter",
    "🌸 Spring": "Spring",
    "☀️ Summer": "Summer",
    "🍂 Autumn": "Autumn",
    "🔥 Final": "Final",
}


def _cell(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return v


def calculate_points(place_value, multiplier=1):
    """Calculate points for a single place value: 1st=30, 2nd=20, 3rd=10, participation=1.
    place_value=0 means participated but didn't place (gets 1 pt).
    place_value=None means didn't participate (gets 0).
    """
    if place_value is None or place_value == "":
        return 0
    try:
        place = int(float(place_value))
    except (ValueError, TypeError):
        return 0
    if place in PLACE_POINTS:
        return PLACE_POINTS[place] * multiplier
    # place=0 or place>3 means participated
    if place >= 0:
        return PARTICIPATION_POINTS * multiplier
    return 0


def calculate_total_points(main_place, extra1, extra2, extra3, multiplier=1):
    """Calculate total points from main place + all extra nominations."""
    total = calculate_points(main_place, multiplier)
    total += calculate_points(extra1, multiplier)
    total += calculate_points(extra2, multiplier)
    total += calculate_points(extra3, multiplier)
    return total


def parse_excel(file_path: str) -> dict:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    data = {"participants": [], "events": {}}

    # Parse participants from "👥 Участники" sheet
    part_sheet = None
    for sn in wb.sheetnames:
        if "Участники" in sn:
            part_sheet = wb[sn]
            break

    # Build participant list and name-by-row map
    participant_names_by_row = {}  # row -> name (from Participants sheet)
    if part_sheet:
        for row in range(4, part_sheet.max_row + 1):
            num = _cell(part_sheet, row, 2)
            name = _cell(part_sheet, row, 3)
            nomination = _cell(part_sheet, row, 4)
            if not name or not nomination:
                continue
            name_str = str(name).strip()
            participant_names_by_row[row] = name_str
            data["participants"].append({
                "num": int(num) if num else row - 3,
                "name": name_str,
                "nomination": str(nomination).strip(),
            })

    # Parse event results
    for sheet_name, event_name in SHEET_MAP.items():
        ws = None
        for sn in wb.sheetnames:
            if event_name in sn:
                ws = wb[sn]
                break
        if not ws:
            continue

        # Find the multiplier for this event
        ev_info = next((e for e in SCHOOL_EVENTS if e["name"] == event_name), None)
        multiplier = ev_info["multiplier"] if ev_info else 1

        results = []
        has_data = False

        for row in range(4, ws.max_row + 1):
            # Name in event sheets is a formula reference to Participants sheet
            # With data_only=True it may return cached value or None
            # Use participant_names_by_row as fallback
            name = _cell(ws, row, 3)
            if not name:
                name = participant_names_by_row.get(row)
            if not name:
                continue

            main_place = _cell(ws, row, 5)
            extra1 = _cell(ws, row, 6)
            extra2 = _cell(ws, row, 7)
            extra3 = _cell(ws, row, 8)

            # Calculate total points including extra nominations
            points = calculate_total_points(main_place, extra1, extra2, extra3, multiplier)
            if points > 0:
                has_data = True

            results.append({
                "name": str(name).strip(),
                "main_place": float(main_place) if main_place is not None and main_place != "" else None,
                "extra_nom1": float(extra1) if extra1 is not None and extra1 != "" else None,
                "extra_nom2": float(extra2) if extra2 is not None and extra2 != "" else None,
                "extra_nom3": float(extra3) if extra3 is not None and extra3 != "" else None,
                "points": points,
            })

        data["events"][event_name] = {
            "results": results,
            "has_data": has_data,
        }

    wb.close()
    return data


async def import_excel_to_db(session: AsyncSession, file_path: str, season_name: str = None):
    data = parse_excel(file_path)

    # Get or create current season
    result = await session.execute(select(Season).where(Season.is_current == True))
    season = result.scalar_one_or_none()

    if not season:
        season = Season(name=season_name or "Сезон 2025/2026", is_current=True)
        session.add(season)
        await session.flush()

    # Clear old data for this season
    await session.execute(delete(Result).where(
        Result.participant_id.in_(
            select(Participant.id).where(Participant.season_id == season.id)
        )
    ))
    await session.execute(delete(Participant).where(Participant.season_id == season.id))

    old_events = await session.execute(
        select(Event).where(Event.season_id == season.id, Event.event_type == "school")
    )
    for ev in old_events.scalars().all():
        await session.execute(delete(Result).where(Result.event_id == ev.id))
        await session.delete(ev)

    await session.flush()

    # Create participants
    participant_map = {}
    for p in data["participants"]:
        part = Participant(name=p["name"], nomination=p["nomination"], season_id=season.id)
        session.add(part)
        await session.flush()
        participant_map[p["name"]] = part

    # Create events and results
    updated_events = []
    for ev_info in SCHOOL_EVENTS:
        ev_name = ev_info["name"]
        ev_data = data["events"].get(ev_name, {"results": [], "has_data": False})

        status = "completed" if ev_data["has_data"] else "upcoming"

        event = Event(
            name=ev_name,
            emoji=ev_info["emoji"],
            event_type="school",
            season_id=season.id,
            status=status,
            multiplier=ev_info["multiplier"],
            sort_order=ev_info["sort_order"],
        )
        session.add(event)
        await session.flush()

        if ev_data["has_data"]:
            updated_events.append(ev_name)

        for r in ev_data["results"]:
            part = participant_map.get(r["name"])
            if not part:
                continue
            res = Result(
                participant_id=part.id,
                event_id=event.id,
                main_place=r["main_place"],
                extra_nom1=r["extra_nom1"],
                extra_nom2=r["extra_nom2"],
                extra_nom3=r["extra_nom3"],
                points=r["points"],
            )
            session.add(res)

    await session.commit()
    return {"season_id": season.id, "participants": len(participant_map), "updated_events": updated_events}
